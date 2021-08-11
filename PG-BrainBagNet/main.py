import os
import GPUtil
import argparse
import setting as st
import setting_2 as fst
from scheduler import GradualWarmupScheduler
import torch
import construct_model
import numpy as np
from torch.backends import cudnn
import utils as ut

from data_load import data_load as DL
from data_load import jsy_data_load as jDL
from test import *
from train import *
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import random

seed = 1234
random.seed(seed)
np.random.seed(seed)
torch.manual_seed(seed)
torch.cuda.manual_seed(seed)
torch.cuda.manual_seed_all(seed)
torch.backends.cudnn.deterministic = True
torch.backends.cudnn.benchmark = False
# torch.backends.cudnn.benchmark = True
# CUDA_LAUNCH_BLOCKING=1

def main(config):
    """GPU connection"""
    if config.gpu is None :
        # devices = "%d" % GPUtil.getFirstAvailable(order="memory")[0]
        devices = 2
        os.environ["CUDA_VISIBLE_DEVICES"] = str(devices)
    else :
        devices = str(config.gpu)
        os.environ["CUDA_VISIBLE_DEVICES"] = devices

    """ 1. data process """
    if fst.flag_orig_npy == True:
        print('preparation of the numpy')
        if os.path.exists(st.orig_npy_dir) == False :
            os.makedirs(st.orig_npy_dir)
        """ processing """
        jDL.Prepare_data_4()

    """ 2. fold index processing """
    if fst.flag_fold_index == True:
        print('preparation of the fold index')
        if os.path.exists(st.fold_index_dir) == False:
            os.makedirs(st.fold_index_dir)
        """ save the fold index """
        ut.preparation_fold_index(config)

    """ start and end fold """
    start_fold = st.start_fold
    end_fold = st.end_fold

    """ workbook for the classication performance by fold """
    list_dir_result = []
    list_wb = []
    list_ws = []
    for i in range(len(st.list_standard_eval_dir)):
        list_dir_result.append(st.dir_to_save_1 + st.list_standard_eval_dir[i])
        ut.make_dir(dir=list_dir_result[i], flag_rm=False)
        out = ut.excel_setting(start_fold=start_fold, end_fold=end_fold, result_dir=list_dir_result[i], f_name='results')
        list_wb.append(out[0])
        list_ws.append(out[1])

    """ Loop for fold """
    list_eval_metric = st.list_eval_metric
    metric_avg = [[[] for j in range(len(st.list_eval_metric))] for i in range(len(st.list_standard_eval_dir))]
    for fold in range(start_fold, end_fold+1):
        print("FOLD : {}".format(fold))

        ## TODO : Directory preparation
        print('-' * 10 + 'Directory preparation' + '-' * 10)

        list_dir_save_model_1 = []
        list_dir_confusion_1 = []
        list_dir_heatmap_1 = []
        for i in range(len(st.list_standard_eval_dir)):
            """ dir to save model """
            list_dir_save_model_1.append(
                st.dir_to_save_1 + st.list_standard_eval_dir[i] + '/weights/fold_{}'.format(fold))
            ut.make_dir(dir=list_dir_save_model_1[i], flag_rm=False)

            list_dir_confusion_1.append(
                st.dir_to_save_1 + st.list_standard_eval_dir[i] + '/confusion/fold_{}'.format(fold))
            ut.make_dir(dir=list_dir_confusion_1[i], flag_rm=False)

            list_dir_heatmap_1.append(
                st.dir_to_save_1 + st.list_standard_eval_dir[i] + '/heatmap/fold_{}'.format(fold))
            ut.make_dir(dir=list_dir_heatmap_1[i], flag_rm=False)

        """ dir to save pyplot """
        dir_pyplot_1 = st.dir_to_save_1 + '/pyplot/fold_{}'.format(fold)
        ut.make_dir(dir=dir_pyplot_1, flag_rm=False)


        """ normal classification tasks """
        list_test_result = []
        print('-' * 10 + 'start training' + '-' * 10)
        """ --------------------------------------- """
        """ ------------ classification------------ """
        """ --------------------------------------- """

        """ model construction """
        print('-' * 10 + 'Model construction' + '-' * 10)

        model_1 = construct_model.construct_model(config, flag_model_num=0)

        """pretrained model """
        if fst.flag_pretrained ==True:
            for i_dir in range(len(st.list_dir_preTrain)):
                dir_to_load = st.list_dir_preTrain[i_dir]
                dir_load_model = dir_to_load + '/weights/fold_{}'.format(fold)
                assert os.path.exists(dir_load_model)

                print('load pretrained')
                print(dir_load_model)
                model_dir = ut.model_dir_to_load(fold, dir_load_model)
                pretrained_dict = torch.load(model_dir)
                model_dict = model_1.state_dict()
                for k, v in pretrained_dict.items():
                    if k in model_dict:
                        print(k)
                pretrained_dict = {k: v for k, v in pretrained_dict.items() if k in model_dict}
                model_dict.update(pretrained_dict)
                model_1.load_state_dict(model_dict)

        """ optimizer """
        optimizer_1 = torch.optim.Adam(model_1.parameters(), lr=st.hyperParam_s1.lr, betas=(0.9, 0.999), eps=1e-8, weight_decay=st.hyperParam_s1.weight_decay)

        """ scheduler """
        scheduler_cosine_1 = torch.optim.lr_scheduler.CosineAnnealingLR(optimizer_1, st.hyperParam_s1.epoch)
        scheduler_1 = GradualWarmupScheduler(optimizer_1, multiplier=1, total_epoch=5, after_scheduler=scheduler_cosine_1)


        """ data loader """
        print('-' * 10 + 'data loader' + '-' * 10)
        train_loader = DL.convert_Dloader_3(fold, list_class=st.list_class_for_train, flag_tr_val_te='train', batch_size=st.hyperParam_s1.batch_size, num_workers=0, shuffle=True, drop_last=True)
        val_loader = DL.convert_Dloader_3(fold, list_class=st.list_class_for_test,  flag_tr_val_te='val', batch_size=st.hyperParam_s1.v_batch_size, num_workers=0, shuffle=False, drop_last=False)
        test_loader = DL.convert_Dloader_3(fold, list_class=st.list_class_for_test,  flag_tr_val_te='test', batch_size=st.hyperParam_s1.v_batch_size, num_workers=0, shuffle=False, drop_last=False)
        dict_data_loader = {'train': train_loader,
                            'val': val_loader,
                            'test': test_loader}

        """ training """
        if fst.flag_training == True:
            train.train(config, fold, model_1, dict_data_loader, optimizer_1, scheduler_1, st.hyperParam_s1, list_dir_save_model_1, list_dir_heatmap_1,  dir_pyplot_1, Validation=True, Test_flag=True)

        """ test classification model """
        for i_tmp in range(len(st.list_standard_eval_dir)):
            dict_test_output = test.test(config, fold, model_1, dict_data_loader['test'], st.hyperParam_s1, list_dir_save_model_1[i_tmp], list_dir_heatmap_1[i_tmp], list_dir_confusion_1[i_tmp])
            list_test_result.append(dict_test_output)

        """ fill out the results """
        for i_standard in range(len(st.list_standard_eval_dir)):
            for i in range(len(list_eval_metric)):
                if list_eval_metric[i] in list_test_result[i_standard]:
                    list_ws[i_standard].cell(row=2 + i + st.push_start_row, column=fold + 1, value="%.4f" % (list_test_result[i_standard][list_eval_metric[i]]))
                    metric_avg[i_standard][i].append(list_test_result[i_standard][list_eval_metric[i]])

            for i in range(len(list_eval_metric)):
                if metric_avg[i_standard][i]:
                    avg = round(np.mean(metric_avg[i_standard][i]), 4)
                    std = round(np.std(metric_avg[i_standard][i]), 4)
                    tmp = "%.4f \u00B1 %.4f" % (avg, std)
                    list_ws[i_standard].cell(row=2 + st.push_start_row + i, column=end_fold + 2, value=tmp)

            n_row = list_ws[i_standard].max_row
            n_col = list_ws[i_standard].max_column
            for i_row in range(1, n_row + 1):
                for i_col in range(1, n_col + 1):
                    ca1 = list_ws[i_standard].cell(row=i_row, column=i_col)
                    ca1.alignment = Alignment(horizontal='center', vertical='center')
            list_wb[i_standard].save(list_dir_result[i_standard] + "/results.xlsx")
            list_wb[i_standard].close()

        del model_1, train_loader, test_loader, optimizer_1
        print("finished (fold)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--kfold', type=int, default=st.kfold)
    parser.add_argument('--num_classes', type=int, default=st.num_class)
    parser.add_argument('--lr', type=float, default=st.hyperParam_s1.lr)
    parser.add_argument('--batch_size',type=int, default=st.hyperParam_s1.batch_size)
    parser.add_argument('--v_batch_size', type=int, default=st.hyperParam_s1.v_batch_size)
    parser.add_argument('--num_epochs', type=int, default=st.hyperParam_s1.epoch)
    parser.add_argument('--selected_model', type=str, default=st.model_name)
    parser.add_argument('--gpu', type=str, default=None)

    config = parser.parse_args()
    main(config)

