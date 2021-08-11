import nibabel as nib
from modules import *
from matplotlib import gridspec
import utils as ut
from torch.nn import functional as F
from torch.autograd import Variable
import torch.nn as nn
import shutil
import numpy as np
import os
import setting as st
import setting_2 as fst
import nibabel as nib
import matplotlib
from matplotlib import cm
from matplotlib.colors import ListedColormap, LinearSegmentedColormap
matplotlib.use('Agg')
import torch
from sklearn import metrics
from sklearn.metrics import confusion_matrix
from sklearn.utils.multiclass import unique_labels
import matplotlib.pyplot as plt
import utils
import pickle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from collections import deque


def plot_training_info_1(fold, dir_pyplot, EMS, hyperParam, flag = 'minmax', flag_match = False):
    """ plot the chat"""
    """ train loss """
    y_list = []
    y_list_name = []
    y_list.append(EMS.train_loss)
    y_list_name.append('train loss')
    ut.plot_list_v1(EMS.train_step, y_list, title='train loss', n_xlabel='step', n_ylabel=y_list_name,
                    save_dir=dir_pyplot, file_name='/fold_{0}_train_loss'.format(fold), flag=flag, flag_match=flag_match)
    del y_list, y_list_name

    """ train aux loss """
    tmp_flag = False
    y_list = []
    y_list_name = []
    for tmp_i in range(st.max_num_loss):
        if len(EMS.train_aux_loss[tmp_i]) == len(EMS.train_step):
            y_list.append(EMS.train_aux_loss[tmp_i])
            # y_list_name.append('aux loss {}'.format(tmp_i + 1))
            y_list_name.append('aux loss {}'.format((tmp_i)))
            tmp_flag = True

    if tmp_flag == True:
        ut.plot_list_v1(EMS.train_step, y_list, title='train aux loss', n_xlabel='step', n_ylabel=y_list_name,
                        save_dir=dir_pyplot, file_name='/fold_{0}_train_aux_loss'.format(fold), flag=flag, flag_match=flag_match)
    del y_list, y_list_name


    """ val test acc """
    tmp_flag = False
    y_list = []
    y_list_name = []
    if len(EMS.dict_val_metric['val_acc']) == len(EMS.val_step):
        y_list.append(EMS.dict_val_metric['val_acc'])
        y_list_name.append('val acc')
        tmp_flag = True
    if len(EMS.test_acc) == len(EMS.val_step):
        y_list.append(EMS.test_acc)
        y_list_name.append('test acc')
        tmp_flag = True
    if tmp_flag == True:
        ut.plot_list_v2(EMS.val_step, y_list, title='Val Test Acc Img', n_xlabel='step', n_ylabel=y_list_name,
                        save_dir=dir_pyplot, file_name='/fold_{0}_val_test_acc'.format(fold), flag=flag)
    del y_list, y_list_name

    """ val test AUC """
    tmp_flag = False
    y_list = []
    y_list_name = []
    if len(EMS.dict_val_metric['val_auc']) == len(EMS.val_step):
        y_list.append(EMS.dict_val_metric['val_auc'])
        y_list_name.append('val AUC')
        tmp_flag = True

    if len(EMS.test_auc) == len(EMS.val_step):
        y_list.append(EMS.test_auc)
        y_list_name.append('test AUC')
        tmp_flag = True
    if tmp_flag == True:
        ut.plot_list_v2(EMS.val_step, y_list, title='Val Test AUC Img', n_xlabel='step', n_ylabel=y_list_name,
                        save_dir=dir_pyplot, file_name='/fold_{0}_val_test_AUC'.format(fold), flag=flag)
    del y_list, y_list_name


    """ val test aux acc"""
    for tmp_i in range(hyperParam.num_aux_cls):
        tmp_flag = False
        y_list = []
        y_list_name = []
        if len(EMS.dict_val_metric['val_acc_aux'][tmp_i]) == len(EMS.val_step):
            y_list.append(EMS.dict_val_metric['val_acc_aux'][tmp_i])
            y_list_name.append('val acc {}'.format(hyperParam.loss_lambda['aux_cls' + hyperParam.name_aux_cls[tmp_i]]))
            tmp_flag = True

        if len(EMS.test_acc_aux[tmp_i]) == len(EMS.val_step):
            y_list.append(EMS.test_acc_aux[tmp_i])
            y_list_name.append('test acc {}'.format(hyperParam.loss_lambda['aux_cls' + hyperParam.name_aux_cls[tmp_i]]))
            tmp_flag = True
        if tmp_flag == True:
            ut.plot_list_v2(EMS.val_step, y_list, title='Val Test Acc {}'.format(tmp_i), n_xlabel='step', n_ylabel=y_list_name,
                            save_dir=dir_pyplot, file_name='/fold_{0}_val_test_acc_{1}'.format(fold, tmp_i), flag=flag)

    """ val test loss """
    tmp_flag = False
    y_list = []
    y_list_name = []
    if len(EMS.dict_val_metric['val_loss']) == len(EMS.val_step):
        y_list.append(EMS.dict_val_metric['val_loss'])
        y_list_name.append('val loss')
        tmp_flag = True
    if len(EMS.test_loss) == len(EMS.val_step):
        y_list.append(EMS.test_loss)
        y_list_name.append('test loss')
        tmp_flag = True
    if tmp_flag == True:
        ut.plot_list_v2(EMS.val_step, y_list, title='Val Test Loss', n_xlabel='step', n_ylabel=y_list_name,
                    save_dir=dir_pyplot, file_name='/fold_{0}_val_test_loss'.format(fold), flag=flag)
    del y_list, y_list_name

    """ learning rate """
    y_list = []
    y_list_name = []
    y_list.append(EMS.LR)
    y_list_name.append('learning rate')
    ut.plot_list_v1(EMS.val_step, y_list, title='Learning rate', n_xlabel='step', n_ylabel=y_list_name,
                    save_dir=dir_pyplot, file_name='/fold_{0}_Learning_rate'.format(fold), flag=flag, flag_match=flag_match)
    del y_list, y_list_name


def smooth_one_hot(targets, n_classes, smoothing = 0.0):
    assert 0 <= smoothing < 1
    with torch.no_grad():
        targets = torch.empty(size = (targets.size(0), n_classes)).cuda().fill_(smoothing / (n_classes-1)).scatter_(1, targets.long().data.unsqueeze(-1), 1-smoothing)
    return targets[:, -1]

class FocalLoss_sigmoid_smooth(nn.Module):
    def __init__(self, gamma=0, smoothing = 0.0, alpha=None, size_average=True):
        super(FocalLoss_sigmoid_smooth, self).__init__()
        self.gamma = gamma
        self.alpha = alpha
        if isinstance(alpha, (float, int)): self.alpha = torch.Tensor([alpha, 1 - alpha])
        if isinstance(alpha,list): self.alpha = torch.Tensor(alpha)
        self.size_average = size_average
        self.smoothing = smoothing

    def forward(self, input, target, reduction=True, logit=True):
        if input.dim()>2:
            input = input.view(input.size(0),input.size(1),-1)  # N,C,H,W => N,C,-1
            input = input.transpose(1,2)    # N,-1, C
            input = input.contiguous().view(-1,input.size(2))   # -1 ,C
        target_2 = smooth_one_hot(target, 2, smoothing=self.smoothing)

        if logit == True:
            pt = torch.sigmoid(input)
        else :
            pt = input

        eps = 1e-7
        term1 = pt ** self.gamma * torch.log(1-pt+eps)
        term2 = (1 - pt) ** self.gamma * torch.log(pt+eps)

        if self.alpha is not None:
            if self.alpha.type()!=input.data.type():
                self.alpha = self.alpha.type_as(input.data).unsqueeze(1)
            term1 = term1 * self.alpha[0]
            term2 = term2 * self.alpha[1]

        loss = -(target_2[:, 0].unsqueeze(-1) * term1 + target_2[:, 1].unsqueeze(-1) * term2)

        if reduction == True:
            if self.size_average:
                return loss.mean()
            else:
                return loss.sum()
        else:
            return loss

def model_save_through_validation(fold, epoch, start_eval_epoch, EMS, selected_EMS, ES, model, dir_save_model, metric_1 = 'val_loss', metric_2=None, save_flag = False):
    """ save the model """
    start_eval_epoch = start_eval_epoch
    # start_eval_epoch = 1

    tmp_flag = False
    if save_flag == False:
        if epoch >= start_eval_epoch:
            if metric_1 == 'val_loss' or metric_1 == 'val_mean_loss':
                ES(EMS.dict_val_metric[metric_1][-1], None)
                if ES.early_stop == False:
                    # loss
                    if selected_EMS.dict_val_metric[metric_1] >= EMS.dict_val_metric[metric_1][-1]:
                        selected_EMS.selected_ep = epoch
                        selected_EMS.dict_val_metric[metric_1] = EMS.dict_val_metric[metric_1][-1]

                        """save model"""
                        if selected_EMS.latest_selceted_model_dir != '':
                            os.remove(selected_EMS.latest_selceted_model_dir)
                        current_model_dir = '%s/fold%d_epoch%d.ckpt' % (dir_save_model, (fold), (epoch))
                        try:
                            torch.save(model.state_dict(), current_model_dir)
                        except KeyboardInterrupt:
                            pass
                        except ValueError:
                            pass
                        selected_EMS.latest_selceted_model_dir = current_model_dir
                        tmp_flag = True
            else:
                ES(None, EMS.dict_val_metric[metric_1][-1])
                if ES.early_stop == False:
                    # accuracy, AUC
                    if selected_EMS.dict_val_metric[metric_1] <= EMS.dict_val_metric[metric_1][-1]:
                        selected_EMS.selected_ep = epoch
                        selected_EMS.dict_val_metric[metric_1] = EMS.dict_val_metric[metric_1][-1]

                        """save model"""
                        if selected_EMS.latest_selceted_model_dir != '':
                            os.remove(selected_EMS.latest_selceted_model_dir)
                        current_model_dir = '%s/fold%d_epoch%d.ckpt' % (dir_save_model, (fold), (epoch))
                        try:
                            torch.save(model.state_dict(), current_model_dir)
                        except KeyboardInterrupt:
                            pass
                        except ValueError:
                            pass
                        selected_EMS.latest_selceted_model_dir = current_model_dir
                        tmp_flag = True
            print('')
            print('------ metric_{} ------'.format(metric_1))
            print('Selected_epoch : {}'.format(selected_EMS.selected_ep))
            print('Selected_val_metric : {}'.format(selected_EMS.dict_val_metric[metric_1]))
            print('')

        else:
            if selected_EMS.latest_selceted_model_dir != '':
                os.remove(selected_EMS.latest_selceted_model_dir)

            current_model_dir = '%s/fold%d_epoch%d.ckpt' % (dir_save_model, (fold), (epoch))
            try:
                torch.save(model.state_dict(), current_model_dir)
            except KeyboardInterrupt:
                pass
            except ValueError:
                pass
            selected_EMS.latest_selceted_model_dir = current_model_dir
            tmp_flag = True
            print('')
            print('------ metric_{} ------'.format(metric_1))
            print('Selected_epoch : {}'.format(selected_EMS.selected_ep))
            print('Selected_val_metric : {}'.format(selected_EMS.dict_val_metric[metric_1]))
            print('')
    else:
        """save model"""
        if selected_EMS.latest_selceted_model_dir_2 != '':
            os.remove(selected_EMS.latest_selceted_model_dir_2)
        current_model_dir = '%s/fold%d_epoch%d.ckpt' % (dir_save_model, (fold), (epoch))
        try:
            torch.save(model.state_dict(), current_model_dir)
        except KeyboardInterrupt:
            pass
        except ValueError:
            pass
        selected_EMS.latest_selceted_model_dir_2 = current_model_dir
        tmp_flag = True
    return tmp_flag


class eval_selected_metirc_storage():
    def __init__(self):
        super(eval_selected_metirc_storage, self).__init__()

        """ saved model info"""
        self.latest_selceted_model_dir = ''
        self.latest_selceted_model_dir_2 = ''
        self.selected_ep = 0
        self.dict_val_metric = {
            'val_mean_loss': 10000,
            'val_loss': 10000,
            'val_acc': 0,
            'val_auc': 0,
        }

class eval_metric_storage():
    def __init__(self):
        super(eval_metric_storage, self).__init__()

        """ learning rate """
        self.LR = []

        """ train """
        self.total_train_iter = 0
        self.total_train_step = 0
        self.train_loss = []

        self.train_aux_loss= [[] for i in range(st.max_num_loss)]
        self.train_acc = []
        self.train_step = []

        """ val """
        self.dict_val_metric = {
            'val_loss_queue': deque([]),
            'val_mean_loss': [],
            'val_loss': [],
            'val_acc': [],
            'val_acc_aux': [[] for i in range(st.max_num_loss)],
            'val_auc': [],
            'val_auc_aux': [[] for i in range(st.max_num_loss)],
            'val_MAE': [],
            'val_loss_age': [],
        }

        self.val_step = []
        self.val_loss_1 = []
        self.val_loss_2 = []
        self.val_loss_3 = []
        self.val_loss_4 = []
        self.val_loss_5 = []

        """ test """
        self.test_loss = []
        self.test_acc = []
        self.test_acc_aux = [[] for i in range(st.max_num_loss)]
        self.test_auc = []
        self.test_auc_aux = [[] for i in range(st.max_num_loss)]
        self.test_MAE = []
        self.test_loss_age = []
        self.test_step = []

        self.test_loss_1 = []
        self.test_loss_2 = []
        self.test_loss_3 = []
        self.test_loss_4 = []
        self.test_loss_5 = []

    def forward(self):
        pass

class EarlyStopping():
    """
    Early Stopping to terminate training early under certain conditions
    """
    def __init__(self, delta=0, patience=5, verbose = True):
        self.delta = delta
        self.patience = patience
        self.verbose = verbose
        self.counter = 0
        self.best_loss = None
        self.best_acc = None
        self.best_mean_loss = None
        self.early_stop = False
        self.val_loss_min = np.Inf
        self.val_acc_max = 0
        self.wait = 0
        self.stopped_epoch = 0
        super(EarlyStopping, self).__init__()

    def __call__(self, val_loss, val_acc):
        if self.early_stop == False:
            if val_loss != None:
                if self.best_loss is None:
                    self.best_loss = val_loss
                # better model has been found.
                if val_loss < self.best_loss + self.delta:
                    self.best_loss = val_loss
                    self.counter = 0
                # saved model is better.
                else:
                    self.counter += 1
                    if self.counter >= self.patience:
                        self.early_stop = True
            else:
                if self.best_acc is None:
                    self.best_acc= val_acc
                # better model has been found.
                if val_acc > self.best_acc + self.delta:
                    self.best_acc = val_acc
                    self.counter = 0
                # saved model is better.
                else:
                    self.counter += 1
                    if self.counter >= self.patience:
                        self.early_stop = True

            if self.verbose == True:
                print(f'Early Stopping counter : {self.counter} out of {self.patience}')
        else:
            pass


def Cross_validation(num_data, k_fold, Random_seed=0):
    indices = np.random.RandomState(seed=Random_seed).permutation(num_data)
    np.random.shuffle(indices)
    num_idx = num_data // k_fold
    sample_remainder = num_data % k_fold
    list_size_each_fold = []
    tmp = 0
    for i_fold in range(k_fold):
        if sample_remainder > i_fold:
            list_size_each_fold.append(num_idx + 1 + tmp)
        else:
            list_size_each_fold.append(num_idx+ tmp)
        tmp = list_size_each_fold[-1]

    train_idx = []
    test_idx = []
    val_idx = []
    for i_fold in range(k_fold):
        fold_slice = np.split(indices.copy(), list_size_each_fold, axis=0)
        fold_slice.pop(-1)
        if i_fold == k_fold - 1:
            test_idx.append(fold_slice.pop(i_fold))
            val_idx.append(fold_slice.pop(0))
        else:
            test_idx.append(fold_slice.pop(i_fold))
            val_idx.append(fold_slice.pop(i_fold))
        train_idx.append(np.concatenate(fold_slice))
    return train_idx, val_idx, test_idx

def search_in_whole_subdir(file_dir, sub_dir, n_file, n_ext='.nii'):
    """
    :param file_dir: file directory
    :param sub_dir: the directory default = ''
    :param n_file: a list which words that extraction included
    :param n_ext: the type of files (e.g., .gt, .nii)
    :return: file list
    """

    """ make dir to save if not exist """
    if os.path.exists(file_dir + sub_dir) == False:
        os.makedirs(file_dir+sub_dir)

    file_list = [] # the list to reture
    for (path, dir, files) in os.walk(file_dir + sub_dir):
        # print(path)
        for filename in files:
            ext = os.path.splitext(filename)[-1] # 0 : filename, 1 : 확장자
            _file = os.path.splitext(filename)[0]
            if ext == n_ext:
                count_ = 0
                for i in range (len(n_file)):
                     if n_file[i] in _file :
                         count_ += 1
                if count_ == len(n_file) :
                    file_to_save = path + '/' + filename
                    file_list.append(file_to_save)
    # print(len(file_list))
    return file_list



def plot_confusion_matrix(y_true, y_pred, classes,f_dir, f_name, normalize=False, title=None, cmap=plt.cm.Blues):
    """
    This function prints and plots the confusion matrix.
    Normalization can be applied by setting `normalize=True`.
    """
    if not title:
        if normalize:
            title = 'Normalized confusion matrix'
        else:
            title = 'Confusion matrix, without normalization'

    # Compute confusion matrix
    cm = confusion_matrix(y_true, y_pred)

    # Only use the labels that appear in the data
    classes = classes[unique_labels(y_true, y_pred)]
    if normalize:
        cm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
        print("Normalized confusion matrix")
    else:
        print('Confusion matrix, without normalization')

    print(cm)

    plt.rcParams.update({'font.size': 10})
    fig, ax = plt.subplots()
    im = ax.imshow(cm, interpolation='nearest', cmap=cmap)
    ax.figure.colorbar(im, ax=ax)

    # We want to show all ticks...
    ax.set(xticks=np.arange(cm.shape[1]),
           yticks=np.arange(cm.shape[0]),
           # ... and label them with the respective list entries
           xticklabels=classes, yticklabels=classes,
           title=title,
           ylabel='True label',
           xlabel='Predicted label')

    # Rotate the tick labels and set their alignment.
    plt.setp(ax.get_xticklabels(), rotation=45, ha="right",
             rotation_mode="anchor")

    # Loop over data dimensions and create text annotations.
    fmt = '.2f' if normalize else 'd'
    thresh = cm.max() / 2.
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            ax.text(j, i, format(cm[i, j], fmt),
                    ha="center", va="center",
                    color="white" if cm[i, j] > thresh else "black")
    fig.tight_layout()
    plt.savefig(f_dir + f_name)
    plt.close(fig)

def preparation_fold_index(config):
    list_trIdx = [] # (# of class, # of fold)
    list_valIdx = []  # (# of class, # of fold)
    list_teIdx = []  # (# of class, # of fold)

    for i in range(len(st.list_class_type)):
        if st.list_data_type[st.data_type_num] == 'Density':
            tmp_shape = \
            np.memmap(filename=st.ADNI_fold_image_path[i], mode="r", dtype=np.float64).reshape(-1, st.num_modality,
                                                                                               st.x_size, st.y_size,
                                                                                               st.z_size).shape[0]
        elif st.list_data_type[st.data_type_num] == 'ADNI_JSY':
            tmp_shape = \
            np.memmap(filename=st.ADNI_fold_image_path[i], mode="r", dtype=np.float32).reshape(-1, st.num_modality,
                                                                                               st.x_size, st.y_size,
                                                                                               st.z_size).shape[0]
        elif st.list_data_type[st.data_type_num] == 'ADNI_JSY_2':
            tmp_shape = \
            np.memmap(filename=st.ADNI_fold_image_path[i], mode="r", dtype=np.float32).reshape(-1, st.num_modality,
                                                                                               st.x_size, st.y_size,
                                                                                               st.z_size).shape[0]
        elif st.list_data_type[st.data_type_num] == 'ADNI_JSY_3':
            tmp_shape = \
            np.memmap(filename=st.ADNI_fold_image_path[i], mode="r", dtype=np.float32).reshape(-1, st.num_modality,
                                                                                               st.x_size, st.y_size,
                                                                                               st.z_size).shape[0]
        elif 'ADNI_Jacob' in st.list_data_type[st.data_type_num] or 'ADNI_AAL_256' in st.list_data_type[st.data_type_num]:
            tmp_shape = \
            np.memmap(filename=st.ADNI_fold_image_path[i], mode="r", dtype=np.uint8).reshape(-1, st.num_modality,
                                                                                               st.x_size, st.y_size,
                                                                                               st.z_size).shape[0]

        print(tmp_shape)
        tmp_trIdx, tmp_valIdx, tmp_teIdx = utils.Cross_validation(tmp_shape, config.kfold, Random_seed=0)
        list_trIdx.append(tmp_trIdx)
        list_valIdx.append(tmp_valIdx)
        list_teIdx.append(tmp_teIdx)

    """ Check whether all of the index is different """
    for i_class_type in range(len(st.list_class_type)):
        for i_fold in range(config.kfold):

            for j_class_type in range(len(st.list_class_type)):
                for j_fold in range(config.kfold):

                    if i_fold != j_fold  or i_class_type != j_class_type:
                        assert not(np.array_equal(list_trIdx[i_class_type][i_fold], list_trIdx[j_class_type][j_fold]))
                        assert not(np.array_equal(list_valIdx[i_class_type][i_fold], list_valIdx[j_class_type][j_fold]))
                        assert not(np.array_equal(list_teIdx[i_class_type][i_fold], list_teIdx[j_class_type][j_fold]))


    """ save index for each class """
    for i_class_type in range(len(st.list_class_type)):
        with open(st.train_index_dir[i_class_type], 'wb') as fp:
            pickle.dump(list_trIdx[i_class_type], fp)
        with open(st.val_index_dir[i_class_type], 'wb') as fp:
            pickle.dump(list_valIdx[i_class_type], fp)
        with open(st.test_index_dir[i_class_type], 'wb') as fp:
            pickle.dump(list_teIdx[i_class_type], fp)



def plot_list_v1(x, y, title ='None', n_xlabel ='x', n_ylabel ='y', save_dir ='', file_name ='', flag ='minmax', flag_match = False):
    np_save_dir = save_dir + '/np'
    if os.path.exists(save_dir) == False:
        os.makedirs(save_dir)
    if os.path.exists(np_save_dir) == False:
        os.makedirs(np_save_dir)
    x_range = [None] * 2
    y_range = [None] * 2
    margin = 0.05

    fig = plt.figure(figsize=(40, 10 * len(y)))
    fig.suptitle(title, fontsize=50)
    plt.rcParams.update({'font.size': 22})

    ##TODO : x_range
    x_range[0] = min(x)
    x_range[1] = max(x)

    ##TODO : y_range
    if flag_match == True:
        if flag == 'minmax':
            y_range[0] = np.vstack(y).min()
            y_range[1] = np.vstack(y).max()

        elif flag == 'percentile':
            y_range[0] = np.percentile(np.vstack(y), 1)
            y_range[1] = np.percentile(np.vstack(y), 99)

        elif flag == 'dist':
            mean = np.vstack(y).mean()
            std = np.vstack(y).std()
            y_range[0] = mean - 5 * std
            y_range[1] = mean + 5 * std

    ##TODO: plotting ans save
    for i in range(len(y)):
        ax1 = fig.add_subplot(len(y), 1, i + 1)
        # ax1.set_title(title + '_{}'.format(i))
        ax1.set_ylabel(n_ylabel[i], color='b')
        ax1.set_xlabel(n_xlabel, color='b')
        ax1.plot(x, y[i], c='b', ls='-', marker='.', label=n_ylabel[i])  # ls : :, -, o-, .-
        plt.grid(True)
        plt.legend()
        if flag_match != True:
            if flag == 'minmax':
                y_range[0] = np.array(y[i]).min()
                y_range[1] = np.array(y[i]).max()

            elif flag == 'percentile':
                y_range[0] = np.percentile(y[i], 1)
                y_range[1] = np.percentile(y[i], 99)

            elif flag == 'dist':
                mean = np.array(y[i]).mean()
                std = np.array(y[i]).std()
                y_range[0] = mean - 5 * std
                y_range[1] = mean + 5 * std
        if (x_range[1] - x_range[0]) > 0:
            plt.xlim(x_range[0] - (x_range[1] - x_range[0]) * margin, x_range[1] + (x_range[1] - x_range[0]) * margin)
        if (y_range[1] - y_range[0]) > 0:
            plt.ylim(y_range[0] - (y_range[1] - y_range[0]) * margin, y_range[1] + (y_range[1] - y_range[0]) * margin)
        np.save(file=np_save_dir + file_name + '_' + n_ylabel[i], arr=y[i])

    plt.savefig(os.path.join(save_dir + file_name))

    plt.close('all')

def plot_list_v2(x, y, title ='None', n_xlabel ='x', n_ylabel ='y', save_dir ='', file_name ='', flag ='minmax'):
    np_save_dir = save_dir + '/np'
    if os.path.exists(save_dir) == False:
        os.makedirs(save_dir)
    if os.path.exists(np_save_dir) == False:
        os.makedirs(np_save_dir)

    x_range = [None] * 2
    y_range = [None] * 2
    margin = 0.05

    fig = plt.figure(figsize=(40, 10))
    fig.suptitle(title, fontsize=50)
    plt.rcParams.update({'font.size': 22})
    ##TODO : x_range
    x_range[0] = min(x)
    x_range[1] = max(x)

    ##TODO : y_range
    if flag == 'minmax':
        y_range[0] = np.vstack(y).min()
        y_range[1] = np.vstack(y).max()

    elif flag == 'percentile':
        y_range[0] = np.percentile(np.vstack(y), 1)
        y_range[1] = np.percentile(np.vstack(y), 99)

    elif flag == 'dist':
        mean = np.vstack(y).mean()
        std = np.vstack(y).std()
        y_range[0] = mean - 5 * std
        y_range[1] = mean + 5 * std

    ##TODO: plotting ans save
    ax1 = fig.add_subplot(1, 1, 1)
    # ax1.set_title(title)
    # ax1.set_ylabel(n_ylabel, color='k')
    ax1.set_xlabel(n_xlabel, color='k')

    list_color = ['b', 'g', 'r', 'c', 'm', 'y', 'k']
    for i in range(len(y)):
        ax1.plot(x, y[i], c=list_color[i], ls='-', marker='.', label=n_ylabel[i])  # ls : :, -, o-, .-
        np.save(file=np_save_dir + file_name +'_' + n_ylabel[i], arr=y[i])

    if (x_range[1] - x_range[0]) > 0:
        plt.xlim(x_range[0] - (x_range[1] - x_range[0]) * margin, x_range[1] + (x_range[1] - x_range[0]) * margin)
    if (y_range[1] - y_range[0]) > 0:
        plt.ylim(y_range[0] - (y_range[1] - y_range[0]) * margin, y_range[1] + (y_range[1] - y_range[0]) * margin)

    plt.grid(True)
    plt.legend()
    plt.savefig(os.path.join(save_dir + file_name))
    plt.close('all')

def model_dir_to_load(fold, model_load_dir):
    """ find the maximum epoch model between saved models"""
    included_file_name = ['fold' + str(fold)]
    # get the model corresponding to the specific fold
    models = search_in_whole_subdir('', model_load_dir, included_file_name, '.ckpt')
    s_index = 0 # start index
    e_index = 0 # end index
    n_epoch = []
    for i in range (len(models)):
        for j in range (len(models[i])):
            if models[i][-(j+1)] == 'h':
                s_index = j
                break

        for j in range (len(models[i])):
            if models[i][-(j+1)] == '.':
                e_index = j
                break
        n_epoch.append(models[i][-(s_index+1)+1 : -(e_index+1)])

    if len(n_epoch) == 0:
        print("There is no selected model!")
        return None
    else:
        included_file_name.append(max(n_epoch))
        # get the model corresponding to the specific max epoch
        models = search_in_whole_subdir('', model_load_dir, included_file_name, '.ckpt')
        model_dir = models[0]
    return model_dir


def smooth_one_hot(true_labels: torch.Tensor, classes: int, smoothing=0.0):
    """
    if smoothing == 0, it's one-hot method
    if 0 < smoothing < 1, it's smooth method

    """
    assert 0 <= smoothing < 1
    confidence = 1.0 - smoothing
    label_shape = torch.Size((true_labels.size(0), classes))
    with torch.no_grad():
        true_dist = torch.empty(size=label_shape, device=true_labels.device)
        true_dist.fill_(smoothing / (classes - 1))
        true_dist.scatter_(1, true_labels.data.unsqueeze(1), confidence)
    return true_dist


def crop_tensor(datas, start_point, width_size):

    return datas[start_point[0] : start_point[0] + width_size[0],
           start_point[1] : start_point[1] + width_size[1],
           start_point[2] : start_point[2] + width_size[2]]

def excel_setting(start_fold, end_fold, result_dir, f_name):
    """ setting for the excel file """
    wb = Workbook()
    ws1 = wb.create_sheet('train_result', 0)
    exp_name = st.exp_title
    exp_description = st.exp_description

    """excel setting"""

    """ first col"""
    ws1.cell(row=1 + st.push_start_row, column=1, value="fold")
    for i in range(len(st.list_eval_metric)):
        ws1.cell(row=2+i + st.push_start_row, column=1, value=st.list_eval_metric[i])

    """ first row"""
    for col in range(start_fold, end_fold + 1):
        ws1.cell(row=1 + st.push_start_row, column=col + 1, value="fold_" + str(col))
    ws1.cell(row=1 + st.push_start_row, column=end_fold + 2, value="Avg")
    ws1.cell(row=1 + st.push_start_row, column=end_fold + 2).font = Font(name='Calibri', size=12, bold=True)
    column = str(chr(64 + end_fold + 2))
    ws1.column_dimensions[column].width = 20

    """ head """
    n_row = ws1.max_row
    n_col = ws1.max_column
    ws1.merge_cells(start_row=1, end_row = 1, start_column= 1, end_column=n_col)
    ws1.merge_cells(start_row=2, end_row = 2, start_column= 1, end_column=n_col)
    ws1.cell(row=1, column=1, value=exp_name)
    ws1.cell(row=2, column=1, value=exp_description)


    box = Border(left=Side(style='thin'),
                 right=Side(style='thin'),
                 top=Side(style='thin'),
                 bottom=Side(style='thin'))

    """save xlsx"""
    n_row = ws1.max_row
    n_col = ws1.max_column
    ws1.column_dimensions['A'].width = 20
    for i_row in range(1, n_row+1):
        for i_col in range(1, n_col+1):
            ca1 = ws1.cell(row = i_row, column = i_col)
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            ca1.border = box
            if i_col == 1:
                ca1.font = Font(name='Calibri', size = 15, bold=True)
    wb.save(result_dir + "/{}.xlsx".format(f_name))
    return wb, ws1



def make_dir(dir = './', flag_rm = False, flag_assert = False):
    if flag_rm == True:
        shutil.rmtree(dir)
        os.makedirs(dir)
    else:
        if os.path.exists(dir) == False:
            os.makedirs(dir)
        else :
            if flag_assert == True:
                assert os.path.exists(dir) == False



def plot_heatmap_with_overlay(orig_img, heatmap_img, save_dir, flag_norm = False, fig_title = 'Heatmap', thresh=0.5, percentile = 1):
    shape = heatmap_img.shape
    list_interval = []  ## [axis, n]
    for j in range(3):
        tmp_list = []
        for i in np.arange(30, 71, 10):
            tmp_list.append(int(np.percentile(np.arange(0, shape[j]), i)))
        list_interval.append(np.hstack(tmp_list))

    # axis_type = ['Sagittal', 'Coronal', 'Axial']
    axis_type = ['Coronal']

    fig = plt.figure(figsize=(1 * 2, 2 * 2))  # [n, axis]

    # plt.rcParams.update({'font.size': 15})
    plt.rcParams.update({'font.size': 15, 'font.style': 'italic'})
    # fig.suptitle(fig_title, fontsize=20)

    # heights = [1] * len(axis_type)
    heights = [2]
    widths = [10] * (2)
    widths.append(1)
    gs = gridspec.GridSpec(nrows=len(heights),  # row
                           ncols=len(widths),
                           height_ratios=heights,
                           width_ratios=widths,
                           hspace=0.0,
                           wspace=0.0,
                           )

    # for orig
    orig_vmax = np.percentile(orig_img, 100 - percentile)
    orig_vmin = np.percentile(orig_img, percentile)
    # print(orig_vmin, orig_vmax)

    # cmap_orig = plt.get_cmap('Greys')
    top = cm.get_cmap('Greys', 256)
    bottom = cm.get_cmap('Greys', 0)

    # newcolors = np.vstack((top(np.linspace(0, 1, 128)),
    #                        bottom(np.linspace(0, 1, 128))))
    newcolors = np.vstack((top(np.linspace(0.0, 1.0, 256))))
    cmap_orig = ListedColormap(newcolors, name='Greys_v2')
    # cmap_orig = plt.get_cmap('jet')

    N=256
    red = np.ones((N, 4))

    red[:, 0] = np.linspace(1, 200 / 256, N) #R
    red[:, 1] = np.linspace(1, 0 / 256, N) #G
    red[:, 2] = np.linspace(1, 0 / 256, N) #B
    cmap_heatmap = ListedColormap(red)

    """ normalize """
    vmax = 0.9
    vmin = 0.1

    if flag_norm == True:
        heatmap_img = heatmap_img * (heatmap_img > 0)
        heatmap_img -= heatmap_img.min()
        heatmap_img /= heatmap_img.max()

    thresh_max = vmax * thresh
    thresh_min = vmin * thresh

    alpha = 0.5
    axes = []

    j = 0
    i = 0
    ax1 = fig.add_subplot(gs[j, i])
    orig_scattering_img = np.asarray(orig_img[int(list_interval[0][0]), :, :])
    heatmap_scattering_img = np.asarray(heatmap_img[int(list_interval[0][0]), :, :])
    orig_scattering_img = np.rot90(orig_scattering_img)
    heatmap_scattering_img = np.rot90(heatmap_scattering_img)
    heatmap_scattering_img[
        (heatmap_scattering_img < thresh_max) * (heatmap_scattering_img > thresh_min)] = np.nan
    ax1.imshow(orig_scattering_img, cmap=cmap_orig, vmin=orig_vmin, vmax=orig_vmax)
    im = ax1.imshow(heatmap_scattering_img, cmap=cmap_heatmap, alpha=alpha, vmin=vmin, vmax=vmax)
    ax1.set_yticks([])
    ax1.set_xticks([])
    ax1.spines['right'].set_visible(False)
    ax1.spines['top'].set_visible(False)
    ax1.spines['bottom'].set_visible(False)
    ax1.spines['left'].set_visible(False)
    axes.append(ax1)
    ax1.axis('off')

    j = 0
    i = 1
    ax1 = fig.add_subplot(gs[j, i])
    orig_scattering_img = np.asarray(orig_img[:, int(list_interval[1][2]), :])
    heatmap_scattering_img = np.asarray(heatmap_img[:, int(list_interval[1][2]), :])
    orig_scattering_img = np.rot90(orig_scattering_img)
    heatmap_scattering_img = np.rot90(heatmap_scattering_img)
    heatmap_scattering_img[
        (heatmap_scattering_img < thresh_max) * (heatmap_scattering_img > thresh_min)] = np.nan
    ax1.imshow(orig_scattering_img, cmap=cmap_orig, vmin=orig_vmin, vmax=orig_vmax)
    im = ax1.imshow(heatmap_scattering_img, cmap=cmap_heatmap, alpha=alpha, vmin=vmin, vmax=vmax)

    ax1.set_yticks([])
    ax1.set_xticks([])
    ax1.spines['right'].set_visible(False)
    ax1.spines['top'].set_visible(False)
    ax1.spines['bottom'].set_visible(False)
    ax1.spines['left'].set_visible(False)
    axes.append(ax1)
    ax1.axis('off')
    del orig_scattering_img, heatmap_scattering_img

    # (left, bottom, width, height)
    # cax = plt.axes([0.90, 0.1, 0.02, 0.8])  # left, bottom, width, height
    # cbar = fig.colorbar(im, ax=axes, extend='both', cax=cax)
    # cbar = fig.colorbar(im, ax=axes, cax=cax)
    # cbar = fig.colorbar(im, ax=axes, extend='both')

    # cbar.set_ticks(np.array((vmin, thresh_min, thresh_max, vmax)))
    # cbar.set_ticklabels(["%.2f" % (vmin), "%.2f" % (thresh_min), "%.2f" % (thresh_max), "%.2f" % (vmax)])
    # cbar.set_ticks(np.array((vmin, vmax)))
    # cbar.set_ticklabels(['NC', 'AD'])

    # plt.subplots_adjust(bottom=0.1, right=0.6, top=0.9, left=0.5)

    plt.tight_layout()
    plt.savefig(save_dir, dpi=500, bbox_inches='tight')
    plt.close('all')

def plot_heatmap_with_overlay_v2(orig_img, heatmap_img, save_dir, flag_norm = False, fig_title = 'Heatmap', thresh=0.5, percentile = 1):
    list_label_Name = []
    for i_tmp in range(len(st.list_class_type)):
        if st.list_class_for_test[i_tmp] == 1:
            list_label_Name.append(st.list_class_type[i_tmp])

    shape = heatmap_img.shape
    list_interval = []  ## [axis, n]
    for j in range(3):
        tmp_list = []
        for i in np.arange(30, 71, 10):
            tmp_list.append(int(np.percentile(np.arange(0, shape[j]), i)))
        list_interval.append(np.hstack(tmp_list))

    # axis_type = ['Sagittal', 'Coronal', 'Axial']
    axis_type = ['Coronal']

    fig = plt.figure(figsize=(1 * 2, 2 * 2))  # [n, axis]
    # plt.rcParams.update({'font.size': 15})
    plt.rcParams.update({'font.size': 15, 'font.style': 'italic'})
    # fig.suptitle(fig_title, fontsize=20)


    # heights = [1] * len(axis_type)
    heights = [2]
    widths = [10] * (2)
    widths.append(1)
    gs = gridspec.GridSpec(nrows=len(heights),  # row
                           ncols=len(widths),
                           height_ratios=heights,
                           width_ratios=widths,
                           hspace=0.0,
                           wspace=0.0,
                           )

    # cmap_orig = plt.get_cmap('Greys')
    # cmap_orig = plt.get_cmap('jet')
    # cmap_orig = plt.get_cmap('Greys')
    top = cm.get_cmap('Greys', 256)
    bottom = cm.get_cmap('Greys', 0)

    # newcolors = np.vstack((top(np.linspace(0, 1, 128)),
    #                        bottom(np.linspace(0, 1, 128))))
    newcolors = np.vstack((top(np.linspace(0, 1.0, 256))))
    cmap_orig = ListedColormap(newcolors, name='Greys_v2')


    N=256
    red = np.ones((N, 4))
    # red[:, 0] = np.linspace(255 / 256, 1, N)
    # red[:, 1] = np.linspace(0 / 256, 1, N)
    # red[:, 2] = np.linspace(65 / 256, 1, N)

    red[:, 0] = np.linspace(1, 200 / 256, N) #R
    red[:, 1] = np.linspace(1, 0 / 256, N) #G
    red[:, 2] = np.linspace(1, 0 / 256, N) #B
    cmap_heatmap = ListedColormap(red)
    # cmap_heatmap = plt.get_cmap('Reds')
    # cmap_heatmap = plt.get_cmap('coolwarm')
    # cmap_heatmap = plt.get_cmap('jet')
    # cmap_heatmap = plt.get_cmap('hot')
    # cmap_heatmap = plt.get_cmap('bwr')

    # for orig
    orig_vmax = np.percentile(orig_img, 100 - percentile)
    orig_vmin = np.percentile(orig_img, percentile)
    # print(orig_vmin, orig_vmax)

    """ normalize """
    vmax = 30.0
    vmin = 0.0

    # if flag_norm == True:
    #     heatmap_img = heatmap_img * (heatmap_img > 0)
    #     heatmap_img -= heatmap_img.min()
    #     heatmap_img /= heatmap_img.max()

    # vmax = np.percentile(heatmap_img, 100-percentile)
    # vmin = np.percentile(heatmap_img, percentile)
    # vmax = heatmap_img.max()
    # vmin = heatmap_img.min()
    # print(heatmap_img.max())
    # print(heatmap_img.min())
    # print(vmin, vmax)

    # if np.abs(vmax) > np.abs(vmin):
    #     vmax = np.abs(vmax)
    #     vmin = -np.abs(vmax)
    # else:
    #     vmax = np.abs(vmin)
    #     vmin = -np.abs(vmin)

    thresh_max = vmax * thresh
    thresh_min = vmin * thresh
    # thresh_max = np.percentile(heatmap_img, 97)
    # thresh_min = np.percentile(heatmap_img, 3)
    # print(thresh_min, thresh_max)
    # if np.abs(thresh_max) < np.abs(thresh_min):
    #     thresh_max = np.abs(thresh_max)
    #     thresh_min = -np.abs(thresh_max)
    # else:
    #     thresh_max = np.abs(thresh_min)
    #     thresh_min = -np.abs(thresh_min)

    alpha = 0.5
    axes = []

    j = 0
    i = 0
    ax1 = fig.add_subplot(gs[j, i])
    orig_scattering_img = np.asarray(orig_img[int(list_interval[0][0]), :, :])
    heatmap_scattering_img = np.asarray(heatmap_img[int(list_interval[0][0]), :, :])
    orig_scattering_img = np.rot90(orig_scattering_img)
    heatmap_scattering_img = np.rot90(heatmap_scattering_img)
    heatmap_scattering_img[
        (heatmap_scattering_img < thresh_max) * (heatmap_scattering_img > thresh_min)] = np.nan
    ax1.imshow(orig_scattering_img, cmap=cmap_orig, vmin=orig_vmin, vmax=orig_vmax)
    im = ax1.imshow(heatmap_scattering_img, cmap=cmap_heatmap, alpha=alpha, vmin=vmin, vmax=vmax)
    ax1.set_yticks([])
    ax1.set_xticks([])
    ax1.spines['right'].set_visible(False)
    ax1.spines['top'].set_visible(False)
    ax1.spines['bottom'].set_visible(False)
    ax1.spines['left'].set_visible(False)
    axes.append(ax1)
    ax1.axis('off')

    j = 0
    i = 1
    ax1 = fig.add_subplot(gs[j, i])
    orig_scattering_img = np.asarray(orig_img[:, int(list_interval[1][2]), :])
    heatmap_scattering_img = np.asarray(heatmap_img[:, int(list_interval[1][2]), :])
    orig_scattering_img = np.rot90(orig_scattering_img)
    heatmap_scattering_img = np.rot90(heatmap_scattering_img)
    heatmap_scattering_img[
        (heatmap_scattering_img < thresh_max) * (heatmap_scattering_img > thresh_min)] = np.nan
    ax1.imshow(orig_scattering_img, cmap=cmap_orig, vmin=orig_vmin, vmax=orig_vmax)
    im = ax1.imshow(heatmap_scattering_img, cmap=cmap_heatmap, alpha=alpha, vmin=vmin, vmax=vmax)

    ax1.set_yticks([])
    ax1.set_xticks([])
    ax1.spines['right'].set_visible(False)
    ax1.spines['top'].set_visible(False)
    ax1.spines['bottom'].set_visible(False)
    ax1.spines['left'].set_visible(False)
    axes.append(ax1)
    ax1.axis('off')
    del orig_scattering_img, heatmap_scattering_img

    # (left, bottom, width, height)
    # cax = plt.axes([0.90, 0.1, 0.02, 0.8]) # left, bottom, width, height
    # cbar = fig.colorbar(im, ax=axes, extend='both', cax=cax)
    # cbar = fig.colorbar(im, ax=axes, cax=cax)

    # cbar.set_ticks(np.array((vmin, thresh_min, thresh_max, vmax)))
    # cbar.set_ticklabels(["%.2f" % (vmin), "%.2f" % (thresh_min), "%.2f" % (thresh_max), "%.2f" % (vmax)])
    # cbar.set_ticks(np.array((vmin, vmax)))
    # cbar.set_ticklabels(['NC', 'AD'], rotation=90)

    # cbar.ax.set_yticklabels([list_label_Name[0], list_label_Name[1]], rotation=-90)
    # cbar.ax.set_yticklabels(['low', 'high'], rotation=-90)
    # cbar.ax.tick_params(labelsize=18)
    # plt.subplots_adjust(bottom=0.1, right=0.6, top=0.9, left=0.5)

    plt.tight_layout()
    # plt.savefig(save_dir, dpi=100, bbox_inches='tight')
    plt.savefig(save_dir, dpi=500, bbox_inches='tight')
    plt.close('all')

def plot_heatmap_with_overlay_v3(orig_img, heatmap_img, save_dir, flag_norm = False, fig_title = 'Heatmap', thresh=0.5, percentile = 1):
    list_label_Name = []
    for i_tmp in range(len(st.list_class_type)):
        if st.list_class_for_test[i_tmp] == 1:
            list_label_Name.append(st.list_class_type[i_tmp])

    shape = heatmap_img.shape
    list_interval = []
    for j in range(3):
        tmp_list = []
        for i in np.arange(30, 71, 10):
            tmp_list.append(int(np.percentile(np.arange(0, shape[j]), i)))
        list_interval.append(np.hstack(tmp_list))

    # axis_type = ['Sagittal', 'Coronal', 'Axial']
    axis_type = ['Coronal']

    fig = plt.figure(figsize=(1 * 2, 2 * 2))  # [n, axis]
    # plt.rcParams.update({'font.size': 15})
    plt.rcParams.update({'font.size': 15, 'font.style': 'italic'})
    # fig.suptitle(fig_title, fontsize=20)

    # heights = [1] * len(axis_type)
    heights = [2]
    widths = [10] * (2)
    widths.append(1)
    gs = gridspec.GridSpec(nrows=len(heights),  # row
                           ncols=len(widths),
                           height_ratios=heights,
                           width_ratios=widths,
                           hspace=0.0,
                           wspace=0.0,
                           )
    # cmap_orig = plt.get_cmap('Greys')
    # cmap_orig = plt.get_cmap('jet')
    # cmap_orig = plt.get_cmap('Greys')
    top = cm.get_cmap('Greys', 256)
    bottom = cm.get_cmap('Greys', 0)

    # newcolors = np.vstack((top(np.linspace(0, 1, 128)),
    #                        bottom(np.linspace(0, 1, 128))))
    newcolors = np.vstack((top(np.linspace(0, 1.0, 256))))
    cmap_orig = ListedColormap(newcolors, name='Greys_v2')


    # cmap_heatmap = plt.get_cmap('Reds')
    # cmap_heatmap = plt.get_cmap('coolwarm')
    # cmap_heatmap = plt.get_cmap('jet')
    # cmap_heatmap = plt.get_cmap('bwr')
    N=256
    red = np.ones((N, 4))
    blue = np.ones((N, 4))
    blue[:, 0] = np.linspace(0 / 256, 1, N)
    blue[:, 1] = np.linspace(0 / 256, 1, N)
    blue[:, 2] = np.linspace(200 / 256, 1, N)
    red[:, 0] = np.linspace(1, 200 / 256, N) #R
    red[:, 1] = np.linspace(1, 0 / 256, N) #G
    red[:, 2] = np.linspace(1, 0 / 256, N) #B

    red_cmp = ListedColormap(red)
    blue_cmp = ListedColormap(blue)
    newcolors2 = np.vstack((blue_cmp(np.linspace(0, 1, 128)),
                            red_cmp(np.linspace(0, 1, 128))))
    cmap_heatmap = ListedColormap(newcolors2, name='double')



    # for orig
    orig_vmax = np.percentile(orig_img, 100 - percentile)
    orig_vmin = np.percentile(orig_img, percentile)
    # print(orig_vmin, orig_vmax)

    """ normalize """
    vmax = 30.0
    vmin = -30.0

    # if flag_norm == True:
    #     heatmap_img = heatmap_img * (heatmap_img > 0)
    #     heatmap_img -= heatmap_img.min()
    #     heatmap_img /= heatmap_img.max()

    # vmax = np.percentile(heatmap_img, 100-percentile)
    # vmin = np.percentile(heatmap_img, percentile)
    # vmax = heatmap_img.max()
    # vmin = heatmap_img.min()
    # print(heatmap_img.max())
    # print(heatmap_img.min())
    # print(vmin, vmax)

    if np.abs(vmax) > np.abs(vmin):
        vmax = np.abs(vmax)
        vmin = -np.abs(vmax)
    else:
        vmax = np.abs(vmin)
        vmin = -np.abs(vmin)

    thresh_max = vmax * thresh
    thresh_min = vmin * thresh
    # thresh_max = np.percentile(heatmap_img, 97)
    # thresh_min = np.percentile(heatmap_img, 3)
    # print(thresh_min, thresh_max)
    # if np.abs(thresh_max) < np.abs(thresh_min):
    #     thresh_max = np.abs(thresh_max)
    #     thresh_min = -np.abs(thresh_max)
    # else:
    #     thresh_max = np.abs(thresh_min)
    #     thresh_min = -np.abs(thresh_min)

    alpha = 0.5
    axes = []

    j = 0
    i = 0
    ax1 = fig.add_subplot(gs[j, i])
    orig_scattering_img = np.asarray(orig_img[int(list_interval[0][0]), :, :])
    heatmap_scattering_img = np.asarray(heatmap_img[int(list_interval[0][0]), :, :])
    orig_scattering_img = np.rot90(orig_scattering_img)
    heatmap_scattering_img = np.rot90(heatmap_scattering_img)
    heatmap_scattering_img[
        (heatmap_scattering_img < thresh_max) * (heatmap_scattering_img > thresh_min)] = np.nan
    ax1.imshow(orig_scattering_img, cmap=cmap_orig, vmin=orig_vmin, vmax=orig_vmax)
    im = ax1.imshow(heatmap_scattering_img, cmap=cmap_heatmap, alpha=alpha, vmin=vmin, vmax=vmax)
    ax1.set_yticks([])
    ax1.set_xticks([])
    ax1.spines['right'].set_visible(False)
    ax1.spines['top'].set_visible(False)
    ax1.spines['bottom'].set_visible(False)
    ax1.spines['left'].set_visible(False)
    axes.append(ax1)
    ax1.axis('off')

    j = 0
    i = 1
    ax1 = fig.add_subplot(gs[j, i])
    orig_scattering_img = np.asarray(orig_img[:, int(list_interval[1][2]), :])
    heatmap_scattering_img = np.asarray(heatmap_img[:, int(list_interval[1][2]), :])
    orig_scattering_img = np.rot90(orig_scattering_img)
    heatmap_scattering_img = np.rot90(heatmap_scattering_img)
    heatmap_scattering_img[
        (heatmap_scattering_img < thresh_max) * (heatmap_scattering_img > thresh_min)] = np.nan
    ax1.imshow(orig_scattering_img, cmap=cmap_orig, vmin=orig_vmin, vmax=orig_vmax)
    im = ax1.imshow(heatmap_scattering_img, cmap=cmap_heatmap, alpha=alpha, vmin=vmin, vmax=vmax)

    ax1.set_yticks([])
    ax1.set_xticks([])
    ax1.spines['right'].set_visible(False)
    ax1.spines['top'].set_visible(False)
    ax1.spines['bottom'].set_visible(False)
    ax1.spines['left'].set_visible(False)
    axes.append(ax1)
    ax1.axis('off')
    del orig_scattering_img, heatmap_scattering_img

    plt.tight_layout()
    # plt.savefig(save_dir, dpi=100, bbox_inches='tight')
    plt.savefig(save_dir, dpi=500, bbox_inches='tight')
    plt.close('all')


def data_augmentation(datas, cur_epoch):
    start_point = None
    flip_flag_list = None
    random_scale_list = None
    translation_list = None

    """ flip """
    if fst.flag_random_flip == True:
        flip_flag_list = np.random.normal(size=datas.shape[0]) > 0
        datas[flip_flag_list] = datas[flip_flag_list].flip(-3)

    if fst.flag_random_scale == True:
        random_scale_list = []
        for i_tmp in range(datas.size(0)):
            scale = np.random.uniform(0.95, 1.05, 3)
            random_scale_list.append(scale)
            tmp = F.interpolate(datas[i_tmp].unsqueeze(0), scale_factor=tuple(scale), mode='trilinear', align_corners=False)
            p3d = (0, 50, 0, 50, 0, 50)
            tmp = F.pad(tmp, p3d, 'replicate', 0)
            datas[i_tmp] = tmp[0, :, :datas.size(-3), : datas.size(-2), :datas.size(-1)]

    if fst.flag_cropping == True:
        """ padding """
        pad_size = (st.crop_pad_size)
        datas = F.pad(datas, pad_size, "constant", 0)

        """ width """
        # img_size = [st.x_size, st.y_size, st.z_size]
        img_size = [*datas.size()[-3:]]
        tmp_size = [a_i - b_i for a_i, b_i in zip(st.max_crop_size, st.min_crop_size)]
        width_size = []
        for i in range(len(tmp_size)):
            if tmp_size[i] == 0:
                width_size.append(st.min_crop_size[i])
            else:
                width_size.append(np.random.randint(tmp_size[i]) + st.min_crop_size[i])

        """ start """
        if fst.flag_cropping_subject_wise == True:
            start_point = [[] for i in range(datas.size(0))]
            for i in range(len(start_point)):
                for j in range(len(img_size)):
                    if fst.flag_cropping_same_as_eval == True:
                        start_point[i].append((img_size[j] - width_size[j])//2)
                    else:
                        if img_size[j] - width_size[j] != 0:
                            start_point[i].append(np.random.randint(img_size[j] - width_size[j]))
                        else:
                            start_point[i].append(0)
            datas_aug = torch.zeros(size=width_size).cuda().float().unsqueeze(0).unsqueeze(0).repeat(datas.size(0), 1, 1, 1, 1)
            for batch_i in range(datas.size(0)):
                datas_aug[batch_i][0] = ut.crop_tensor(datas[batch_i][0], start_point[batch_i], width_size)
            datas = datas_aug
            del datas_aug
        else:
            start_point = [[] for i in range(1)]
            for i in range(len(start_point)):
                for j in range(len(img_size)):
                    start_point[i].append(np.random.randint(img_size[j] - width_size[j]))

            datas_aug = torch.zeros(size=width_size).cuda().float().unsqueeze(0).unsqueeze(0).repeat(datas.size(0), 1,
                                                                                                     1, 1, 1)
            for batch_i in range(datas.size(0)):
                datas_aug[batch_i][0] = ut.crop_tensor(datas[batch_i][0], start_point[0], width_size)
            datas = datas_aug
            del datas_aug

    """ gaussain noise """
    if fst.flag_Gaussian_noise == True:
        Gaussian_dist = torch.distributions.normal.Normal(loc=torch.tensor([0.0]), scale=torch.tensor([0.01]))
        Gaussian_noise = Gaussian_dist.sample(datas.size()).squeeze(-1)
        datas = datas + Gaussian_noise.cuda()

    if fst.flag_Avgpool == True:
        datas = F.avg_pool3d(datas, kernel_size = 2, stride =2)

    if fst.flag_Gaussian_blur == True:
        datas = ut.GaussianSmoothing(1, 3, torch.FloatTensor(1).uniform_(0, 1.5))(datas)

    dict_result = {
        "datas": datas,
        "info_flip": flip_flag_list, # ([T, F, F, T])
        "info_scale": random_scale_list,
        "info_crop": np.asarray(start_point),
    }
    return dict_result



def train_classification_model(config, fold, epoch, EMS, loader, model, optimizer, hyperParam):
    if st.flag_beta == True:
        alpha = loader.dataset.n_sample[1] / sum(loader.dataset.n_sample)
    else :
        alpha = None

    criterion_cls = ut.FocalLoss_sigmoid_smooth(gamma=st.focal_gamma, smoothing = st.label_smoothing, alpha=alpha, size_average=True)

    """ loader"""
    model.eval()
    loss_tmp = [0] * st.max_num_loss
    loss_tmp_total = 0

    model.train()
    eps = 1e-7
    for i, data_batch in enumerate(loader):
        with torch.no_grad():
            """ input"""
            datas = Variable(data_batch['data'].float()).cuda()
            labels = Variable(data_batch['label'].long()).cuda()

            """ data norm"""
            tmp_datas = datas.view(datas.size(0), -1)
            tmp_mean = tmp_datas.mean(1, keepdim=True)
            tmp_std = tmp_datas.std(1, keepdim=True)
            tmp_datas = (tmp_datas - tmp_mean) / (tmp_std)
            datas = tmp_datas.view_as(datas)

            """ data augmentation """
            dict_result = ut.data_augmentation(datas=datas, cur_epoch=epoch)
            datas = dict_result['datas']
            info_flip = dict_result['info_flip']
            info_scale = dict_result['info_scale']
            info_crop = dict_result['info_crop']


        """ forward propagation """
        dict_result = model(datas, [info_flip, info_scale, info_crop])

        loss_list_1 = []
        count_loss = 0

        """ classification loss """
        output_1 = dict_result['logits']
        s_labels = labels
        loss_2 = criterion_cls(output_1, s_labels) * hyperParam.loss_lambda['cls']
        loss_list_1.append(loss_2)
        loss_tmp[count_loss] += loss_2.data.cpu().numpy()
        if (EMS.total_train_iter + 1) % hyperParam.iter_to_update == 0:
            EMS.train_aux_loss[count_loss].append(loss_tmp[count_loss])
            loss_tmp[count_loss] = 0
        count_loss += 1

        """ entropy loss """
        output_1 = dict_result['entropy']
        for tmp_i in range(len(output_1)):
            term1 = output_1[tmp_i] * torch.log(output_1[tmp_i] + eps)
            term2 = (1- output_1[tmp_i]) * torch.log(1 - output_1[tmp_i] + eps)
            entropy = -(term1 + term2).mean(dim=(2, 3, 4)).mean()
            loss_2 = -entropy * hyperParam.loss_lambda['entropy']

            loss_list_1.append(loss_2)
            loss_tmp[count_loss] += loss_2.data.cpu().numpy()
            if (EMS.total_train_iter + 1) % hyperParam.iter_to_update == 0:
                EMS.train_aux_loss[count_loss].append(loss_tmp[count_loss])
                loss_tmp[count_loss] = 0
            count_loss += 1

        """ add the loss """
        loss = sum(loss_list_1)
        loss.backward()
        loss_tmp_total += loss.data.cpu().numpy()

        """ optimize the model param """
        if (EMS.total_train_iter + 1) % hyperParam.iter_to_update == 0:
            optimizer.step()
            optimizer.zero_grad()

            """ pyplot """
            EMS.total_train_step += 1
            EMS.train_step.append(EMS.total_train_step)
            EMS.train_loss.append(loss_tmp_total)

            """ print the train loss and tensorboard"""
            if (EMS.total_train_step) % 10 == 0:
                print('Epoch [%d/%d], Step [%d/%d],  Loss: %.4f' % (
                epoch, hyperParam.epoch, (i + 1), (len(loader.dataset) // (hyperParam.batch_size)), loss_tmp_total))
            loss_tmp_total = 0
        EMS.total_train_iter += 1

    return model, optimizer, EMS


def eval_classification_model(config, fold, epoch, loader, model, hyperParam, plot_flag = False, confusion_save_dir = None):
    """ loader"""
    if st.flag_beta == True:
        alpha = loader.dataset.n_sample[1] / sum(loader.dataset.n_sample)
    else :
        alpha = None

    criterion_cls = ut.FocalLoss_sigmoid_smooth(gamma=st.focal_gamma, smoothing=0, alpha=alpha, size_average=True)
    correct = 0
    loss_np = 0
    model.eval()

    pos_label = 1
    neg_label = 0

    groundTruth_result = []
    predict_score_result = []
    predict_result = []

    count = 0

    list_label_Name = []
    for i_tmp in range(len(st.list_class_type)):
        if st.list_class_for_test[i_tmp] == 1:
            list_label_Name.append(st.list_class_type[i_tmp])

    n_to_plot = 3
    count_sample = 0
    dict_plot_info_1 = {
        'count_{}_{}'.format(list_label_Name[0], list_label_Name[0]): 0,
        'count_{}_{}'.format(list_label_Name[0], list_label_Name[1]): 0,
        'count_{}_{}'.format(list_label_Name[1], list_label_Name[0]): 0,
        'count_{}_{}'.format(list_label_Name[1], list_label_Name[1]): 0,
    }
    dict_plot_info_2 = {
        'count_{}_{}'.format(list_label_Name[0], list_label_Name[0]): 0,
        'count_{}_{}'.format(list_label_Name[0], list_label_Name[1]): 0,
        'count_{}_{}'.format(list_label_Name[1], list_label_Name[0]): 0,
        'count_{}_{}'.format(list_label_Name[1], list_label_Name[1]): 0,
    }
    list_classEvidence = [[[] for j in range(1)] for k in range(st.num_class)]


    if plot_flag==True and epoch % st.flag_plot_epoch == 0:
        dir_to_save_plot = './plot_Result_4/fold_{}/epoch_{}'.format(fold, epoch)
        ut.make_dir(dir=dir_to_save_plot, flag_rm=False)
        mni_template_dir = st.orig_data_dir + '/mni_masked_cropped.nii.gz'
        mni_template_img = nib.load(mni_template_dir).get_data()

    if fst.flag_cropping == True:
        tmp_size_x_1 = (st.x_size - st.eval_crop_size[0]) // 2
        tmp_size_x_2 = tmp_size_x_1 + st.eval_crop_size[0]
        tmp_size_y_1 = (st.y_size - st.eval_crop_size[1]) // 2
        tmp_size_y_2 = tmp_size_y_1 + st.eval_crop_size[1]
        tmp_size_z_1 = (st.z_size - st.eval_crop_size[2]) // 2
        tmp_size_z_2 = tmp_size_z_1 + st.eval_crop_size[2]

    with torch.no_grad():
        for data_batch in loader:
            count +=1

            """ input"""
            datas = Variable(data_batch['data'].float()).cuda()
            labels = Variable(data_batch['label'].long()).cuda()

            """ normalization """
            tmp_datas = datas.reshape(datas.size(0), -1)
            tmp_mean = tmp_datas.mean(1, keepdim=True)
            tmp_std = tmp_datas.std(1, keepdim=True)
            tmp_datas = (tmp_datas - tmp_mean) / (tmp_std)
            datas = tmp_datas.view_as(datas)

            if fst.flag_cropping == True:
                datas = Variable(datas[:, :,
                                 tmp_size_x_1 : tmp_size_x_2,
                                 tmp_size_y_1 : tmp_size_y_2,
                                 tmp_size_z_1 : tmp_size_z_2].float()).cuda()

            """ forward propagation """
            dict_result = model(datas, None)
            output_1 = dict_result['logits']
            if plot_flag==True and epoch % st.flag_plot_epoch == 0:
                prob = torch.sigmoid(output_1)
                pred = (prob >= torch.FloatTensor([0.5]).cuda()).float()

                for i_img in range(datas.size(0)):
                    """ count sample """
                    count_sample += 1

                    if (pred[i_img][0] == labels[i_img]) == True and labels[i_img] == 0:
                        tmp_name = '{}_{}'.format(list_label_Name[0], list_label_Name[0])
                        name_label = list_label_Name[0]
                        name_pred = list_label_Name[0]
                    elif (pred[i_img][0] != labels[i_img]) == True and labels[i_img] == 0:
                        tmp_name = '{}_{}'.format(list_label_Name[0], list_label_Name[1])
                        name_label = list_label_Name[0]
                        name_pred = list_label_Name[1]
                    elif (pred[i_img][0] != labels[i_img]) == True and labels[i_img] == 1:
                        tmp_name = '{}_{}'.format(list_label_Name[1], list_label_Name[0])
                        name_label = list_label_Name[1]
                        name_pred = list_label_Name[0]
                    else:
                        tmp_name = '{}_{}'.format(list_label_Name[1], list_label_Name[1])
                        name_label = list_label_Name[1]
                        name_pred = list_label_Name[1]

                    list_plot_1 = dict_result['list_plot_1']
                    if dict_plot_info_1['count_{}'.format(tmp_name)] < n_to_plot and list_plot_1 is not None:
                        dict_plot_info_1['count_{}'.format(tmp_name)] += 1
                        dir_to_save_plot_2 = dir_to_save_plot + '/{}/sub_{}'.format(tmp_name, dict_plot_info_1['count_{}'.format(tmp_name)])
                        ut.make_dir(dir=dir_to_save_plot_2, flag_rm=False)
                        for i_plot, i_tensor in enumerate(list_plot_1):
                            tmp_img = F.interpolate(i_tensor, size=[(i_tensor.size()[-3:][i] - 1)* 8 + 1 for i in range(3)], mode='trilinear', align_corners=True)
                            diffX = datas.size()[-3] - tmp_img.size()[-3]
                            diffY = datas.size()[-2] - tmp_img.size()[-2]
                            diffZ = datas.size()[-1] - tmp_img.size()[-1]
                            tmp_img = F.pad(tmp_img, [0, diffZ, 0, diffY, 0, diffX])

                            if pred[i_img][0] == 0:
                                tmp_prob = 1- prob[i_img, 0]
                            else:
                                tmp_prob = prob[i_img, 0]
                            ut.plot_heatmap_with_overlay(orig_img=mni_template_img,
                                                         heatmap_img=tmp_img[i_img, 0].cpu().data.numpy(),
                                                         save_dir= dir_to_save_plot_2 + '/gate_{}_index_{}'.format(i_plot, count_sample),
                                                         flag_norm = False,
                                                         fig_title='gate generated by coord',
                                                         thresh=0.0,
                                                         percentile=0.01)

                    list_plot_2 = dict_result['list_plot_2']
                    if dict_plot_info_2['count_{}'.format(tmp_name)] < n_to_plot and list_plot_2 is not None:
                        dict_plot_info_2['count_{}'.format(tmp_name)] += 1
                        dir_to_save_plot_2 = dir_to_save_plot + '/{}/sub_{}'.format(tmp_name, dict_plot_info_2['count_{}'.format(tmp_name)])
                        ut.make_dir(dir=dir_to_save_plot_2, flag_rm=False)
                        for i_plot, i_tensor in enumerate(list_plot_2):

                            tmp_img = F.interpolate(i_tensor, size=[(i_tensor.size()[-3:][i] - 1)* 8 + 1 for i in range(3)], mode='trilinear', align_corners=True)
                            diffX = datas.size()[-3] - tmp_img.size()[-3]
                            diffY = datas.size()[-2] - tmp_img.size()[-2]
                            diffZ = datas.size()[-1] - tmp_img.size()[-1]
                            tmp_img = F.pad(tmp_img, [0, diffZ, 0, diffY, 0, diffX])
                            # tmp_img = F.interpolate(i_tensor, size=datas.size()[-3:], mode='trilinear')

                            for i_channel in range(i_tensor.shape[1]):

                                ut.plot_heatmap_with_overlay_v3(orig_img=datas[i_img, 0].cpu().data.numpy(),
                                                                heatmap_img=tmp_img[i_img, i_channel].cpu().data.numpy(),
                                                                save_dir=dir_to_save_plot_2 + '/CE_ch{}_{}_{}'.format(i_channel, i_plot, count_sample),
                                                                flag_norm=False,
                                                                fig_title='Label : {0},   Prediction : {1} ({2:04.4f})'.format(name_label, name_pred, tmp_prob),
                                                                thresh=0.0,
                                                                percentile=0.01)

                                if pred[i_img][0] == 0:
                                    tmp_prob = 1 - prob[i_img, 0]
                                    tmp_img = tmp_img * -1.0
                                else:
                                    tmp_prob = prob[i_img, 0]
                                tmp_img = torch.relu(tmp_img)
                                ut.plot_heatmap_with_overlay_v2(orig_img=datas[i_img, 0].cpu().data.numpy(),
                                                                heatmap_img=tmp_img[i_img, i_channel].cpu().data.numpy(),
                                                                save_dir=dir_to_save_plot_2 + '/CE_ReLU_ch{}_{}_{}'.format(i_channel, i_plot, count_sample),
                                                                flag_norm = False,
                                                                fig_title='Label : {0},   Prediction : {1} ({2:04.4f})'.format(name_label, name_pred, tmp_prob),
                                                                thresh=0.0,
                                                                percentile=0.01)
                            if i_tensor.shape[1] == 2:
                                list_classEvidence[labels[i_img]][0].append(tmp_img[i_img, labels[i_img]].cpu().data.numpy())
                            else:
                                list_classEvidence[labels[i_img]][0].append(tmp_img[i_img, 0].cpu().data.numpy())

            loss_list_1 = []
            count_loss = 0
            s_labels = labels
            loss_2 = criterion_cls(output_1, s_labels) * hyperParam.loss_lambda['cls']
            loss_list_1.append(loss_2)
            count_loss += 1

            """ add the loss """
            loss = sum(loss_list_1)
            loss_np += loss.data.cpu().numpy() * output_1.size(0)

            """ stack ground truth and prediction """
            groundTruth_result.append(np.hstack(labels.data.cpu().numpy()).squeeze())

            """ count the correct prediction """
            prob = torch.sigmoid(output_1)
            pred = (prob >= torch.FloatTensor([0.5]).cuda()).float()
            correct += pred.eq(labels.view_as(pred)).sum().cpu().item()
            predict_result.append(pred.data.cpu().numpy().squeeze())
            predict_score_result.append(prob[:, 0].data.cpu().numpy().squeeze())


    """ stacking """
    groundTruth_result = np.hstack(groundTruth_result)
    predict_result = np.hstack(predict_result)
    predict_score_result = np.hstack(predict_score_result)

    """ plot the confusion matrix """
    if confusion_save_dir != None:
        """ plot and save the confusion matrix """
        class_names = np.zeros(len(st.list_selected_for_train), dtype="U10")
        for i in range (len(st.list_selected_for_train)):
            class_names[i] = st.list_selected_for_train[i]

        ut.plot_confusion_matrix(groundTruth_result, predict_result, classes=class_names, f_dir=confusion_save_dir,
                                 f_name='/confusion_fold_{}.png'.format(fold),
                                 title='Confusion matrix, without normalization')

        fpr, tpr, thresholds = metrics.roc_curve(groundTruth_result, predict_score_result, pos_label=pos_label)
        AUC = metrics.auc(fpr, tpr)

        plt.rcParams.update({'font.size': 10})
        fig, ax = plt.subplots()
        ax.plot([0, 1], [0, 1], 'k--')
        ax.plot(fpr, tpr)

        plt.xlabel('False positive rate')
        plt.ylabel('True positive rate')

        plt.title('ROC curve')
        # plt.legend(loc='best')
        plt.savefig(confusion_save_dir + '/ROC_fold_{}.png'.format(fold))
        plt.close(fig)

    """ Confusion matrix , Accuracy, sensitvity and specificity """
    cm1 = confusion_matrix(groundTruth_result, predict_result)

    """ calculate the accuracy """
    total1 = sum(sum(cm1))
    accuracy1 = (cm1[neg_label, neg_label] + cm1[pos_label, pos_label]) / total1
    specificity1 = cm1[neg_label, neg_label] / (cm1[neg_label, neg_label] + cm1[neg_label, pos_label])
    sensitivity1 = cm1[pos_label, pos_label] / (cm1[pos_label, neg_label] + cm1[pos_label, pos_label])
    fpr, tpr, thresholds = metrics.roc_curve(groundTruth_result, predict_score_result, pos_label=pos_label)
    AUC = metrics.auc(fpr, tpr)

    """ calculate the accuracy and print """
    total = len(loader.dataset)
    loss_np = loss_np / total

    dict_result = {
        "Loss": loss_np,
        "Acc": accuracy1,
        "Acc_aux": None,
        "Sen": sensitivity1,
        "Spe": specificity1,
        "AUC": AUC,
    }
    return dict_result


